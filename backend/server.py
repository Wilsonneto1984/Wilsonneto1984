"""Employee Management System - Main Server
Multi-tenant system for employee attendance management
"""
from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path
import os
import logging
from typing import List, Optional, Literal
from datetime import datetime, timezone, timedelta
from uuid import uuid4
import bcrypt
import jwt
import csv
import io
from models import (
    Company, CompanyCreate, CompanyUpdate,
    User, UserCreate, UserLogin, TokenResponse,
    Employee, EmployeeCreate, EmployeeUpdate,
    Shift, ShiftCreate, ShiftUpdate,
    Attendance, AttendanceCreate, AttendanceUpdate
)

# Load environment
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'employee_management_system')]

# JWT Configuration
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

# Create app
app = FastAPI(title="Employee Management System API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ============ AUTH HELPERS ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))


def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token expired")
    except jwt.JWTError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")


def require_role(allowed_roles: List[str]):
    async def role_checker(current_user: dict = Depends(get_current_user)):
        if current_user['role'] not in allowed_roles:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Access denied. Required roles: {allowed_roles}"
            )
        return current_user
    return role_checker


# ============ STARTUP - CREATE SUPER ADMIN ============

@app.on_event("startup")
async def create_super_admin():
    """Create default super admin if not exists"""
    existing = await db.users.find_one({"role": "super_admin"})
    if not existing:
        admin = User(
            email="admin@system.com",
            name="Super Admin",
            role="super_admin",
            company_id=None
        )
        admin_dict = admin.model_dump()
        admin_dict['password'] = hash_password("admin123")
        admin_dict['created_at'] = admin_dict['created_at'].isoformat()
        await db.users.insert_one(admin_dict)
        logger.info("Super admin created: admin@system.com / admin123")


# ============ AUTH ROUTES ============

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    """Login endpoint for all user types"""
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user['password']):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Check company access if not super admin
    if user['role'] != 'super_admin':
        # Validar código da empresa
        if not credentials.company_code:
            raise HTTPException(status_code=400, detail="Company code is required")
        
        company = await db.companies.find_one({"id": user['company_id']}, {"_id": 0})
        if not company:
            raise HTTPException(status_code=404, detail="Company not found")
        
        # Verificar se código da empresa está correto
        if company.get('company_code') != credentials.company_code:
            raise HTTPException(status_code=401, detail="Invalid company code")
        
        if not company.get('active', True):
            raise HTTPException(status_code=403, detail="Company account is inactive")
        
        # Check subscription expiry
        if company.get('subscription_expires_at'):
            expires = datetime.fromisoformat(company['subscription_expires_at']) if isinstance(company['subscription_expires_at'], str) else company['subscription_expires_at']
            if expires < datetime.now(timezone.utc):
                raise HTTPException(status_code=403, detail="Company subscription has expired")
    
    access_token = create_access_token({"sub": user['id']})
    
    if isinstance(user['created_at'], str):
        user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    user_obj = User(**{k: v for k, v in user.items() if k != 'password'})
    return TokenResponse(access_token=access_token, user=user_obj)


@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: dict = Depends(get_current_user)):
    """Get current user information"""
    if isinstance(current_user['created_at'], str):
        current_user['created_at'] = datetime.fromisoformat(current_user['created_at'])
    return User(**{k: v for k, v in current_user.items() if k != 'password'})


# ============ COMPANY ROUTES (Super Admin Only) ============

@api_router.post("/companies", response_model=Company)
async def create_company(
    company_data: CompanyCreate,
    current_user: dict = Depends(require_role(["super_admin"]))
):
    """Create a new company with admin user"""
    existing = await db.companies.find_one({"name": company_data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Company name already exists")
    
    # Verificar se código da empresa já existe
    existing_code = await db.companies.find_one({"company_code": company_data.company_code})
    if existing_code:
        raise HTTPException(status_code=400, detail="Company code already exists")
    
    existing_user = await db.users.find_one({"email": company_data.admin_email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Admin email already registered")
    
    company = Company(
        name=company_data.name,
        company_code=company_data.company_code,
        subscription_type=company_data.subscription_type
    )
    
    if company_data.subscription_type == "monthly":
        company.subscription_expires_at = datetime.now(timezone.utc) + timedelta(days=30)
    else:
        company.subscription_expires_at = datetime.now(timezone.utc) + timedelta(days=365)
    
    company_dict = company.model_dump()
    company_dict['created_at'] = company_dict['created_at'].isoformat()
    company_dict['subscription_expires_at'] = company_dict['subscription_expires_at'].isoformat()
    
    await db.companies.insert_one(company_dict)
    
    admin_user = User(
        email=company_data.admin_email,
        name=company_data.admin_name,
        role="company_admin",
        company_id=company.id
    )
    
    admin_dict = admin_user.model_dump()
    admin_dict['password'] = hash_password(company_data.admin_password)
    admin_dict['created_at'] = admin_dict['created_at'].isoformat()
    
    await db.users.insert_one(admin_dict)
    
    return company


@api_router.get("/companies", response_model=List[Company])
async def get_companies(
    current_user: dict = Depends(require_role(["super_admin"]))
):
    """Get all companies"""
    companies = await db.companies.find({}, {"_id": 0}).to_list(1000)
    for company in companies:
        if isinstance(company.get('created_at'), str):
            company['created_at'] = datetime.fromisoformat(company['created_at'])
        if isinstance(company.get('subscription_expires_at'), str):
            company['subscription_expires_at'] = datetime.fromisoformat(company['subscription_expires_at'])
    return companies


@api_router.get("/companies/{company_id}", response_model=Company)
async def get_company(
    company_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get company details"""
    if current_user['role'] != 'super_admin' and current_user['company_id'] != company_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    if isinstance(company.get('created_at'), str):
        company['created_at'] = datetime.fromisoformat(company['created_at'])
    if isinstance(company.get('subscription_expires_at'), str):
        company['subscription_expires_at'] = datetime.fromisoformat(company['subscription_expires_at'])
    
    return Company(**company)


@api_router.put("/companies/{company_id}", response_model=Company)
async def update_company(
    company_id: str,
    update_data: CompanyUpdate,
    current_user: dict = Depends(require_role(["super_admin"]))
):
    """Update company information"""
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    if update_dict:
        await db.companies.update_one({"id": company_id}, {"$set": update_dict})
    
    updated_company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if isinstance(updated_company.get('created_at'), str):
        updated_company['created_at'] = datetime.fromisoformat(updated_company['created_at'])
    if isinstance(updated_company.get('subscription_expires_at'), str):
        updated_company['subscription_expires_at'] = datetime.fromisoformat(updated_company['subscription_expires_at'])
    
    return Company(**updated_company)


@api_router.post("/companies/{company_id}/logo")
async def upload_company_logo(
    company_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload company logo"""
    if current_user['role'] != 'super_admin' and current_user['company_id'] != company_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot upload logos")
    
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    uploads_dir = ROOT_DIR / "uploads" / "logos"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    
    file_extension = file.filename.split('.')[-1]
    filename = f"{company_id}.{file_extension}"
    file_path = uploads_dir / filename
    
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    logo_url = f"/api/companies/{company_id}/logo/{filename}"
    await db.companies.update_one(
        {"id": company_id},
        {"$set": {"logo_url": logo_url}}
    )
    
    return {"logo_url": logo_url, "message": "Logo uploaded successfully"}


@api_router.get("/companies/{company_id}/logo/{filename}")
async def get_company_logo(company_id: str, filename: str):
    """Get company logo file"""
    file_path = ROOT_DIR / "uploads" / "logos" / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Logo not found")
    return FileResponse(file_path)


# ============ USER ROUTES ============

@api_router.post("/users", response_model=User)
async def create_user(
    user_data: UserCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new user (company_admin can create viewers)"""
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot create users")
    
    if current_user['role'] == 'company_admin':
        if user_data.role != 'company_viewer':
            raise HTTPException(status_code=403, detail="You can only create viewer users")
        if user_data.company_id != current_user['company_id']:
            raise HTTPException(status_code=403, detail="You can only create users for your company")
    
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(
        email=user_data.email,
        name=user_data.name,
        role=user_data.role,
        company_id=user_data.company_id
    )
    
    user_dict = user.model_dump()
    user_dict['password'] = hash_password(user_data.password)
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    
    return user


@api_router.get("/users", response_model=List[User])
async def get_users(
    company_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get users (super_admin sees all, others see their company)"""
    query = {}
    
    if current_user['role'] == 'super_admin':
        if company_id:
            query['company_id'] = company_id
    else:
        query['company_id'] = current_user['company_id']
    
    users = await db.users.find(query, {"_id": 0, "password": 0}).to_list(1000)
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    return users


# ============ EMPLOYEE ROUTES ============

@api_router.post("/employees", response_model=Employee)
async def create_employee(
    employee_data: EmployeeCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new employee"""
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot create employees")
    
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    existing = await db.employees.find_one({
        "company_id": company_id,
        "chapa": employee_data.chapa
    })
    if existing:
        raise HTTPException(status_code=400, detail="Employee with this chapa already exists")
    
    employee = Employee(
        company_id=company_id,
        **employee_data.model_dump()
    )
    
    employee_dict = employee.model_dump()
    employee_dict['created_at'] = employee_dict['created_at'].isoformat()
    
    await db.employees.insert_one(employee_dict)
    
    return employee


@api_router.get("/employees", response_model=List[Employee])
async def get_employees(
    active_only: bool = True,
    turno: Optional[str] = None,
    mo: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get employees"""
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    query = {"company_id": company_id}
    if active_only:
        query["active"] = True
    if turno:
        query["turno"] = turno
    if mo:
        query["mo"] = mo
    
    employees = await db.employees.find(query, {"_id": 0}).to_list(10000)
    for emp in employees:
        if isinstance(emp.get('created_at'), str):
            emp['created_at'] = datetime.fromisoformat(emp['created_at'])
    
    return employees


@api_router.get("/employees/{employee_chapa}", response_model=Employee)
async def get_employee(
    employee_chapa: str,
    current_user: dict = Depends(get_current_user)
):
    """Get employee by chapa"""
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    employee = await db.employees.find_one({
        "company_id": company_id,
        "chapa": employee_chapa
    }, {"_id": 0})
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    if isinstance(employee.get('created_at'), str):
        employee['created_at'] = datetime.fromisoformat(employee['created_at'])
    
    return Employee(**employee)


@api_router.put("/employees/{employee_chapa}", response_model=Employee)
async def update_employee(
    employee_chapa: str,
    employee_data: EmployeeUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update employee information"""
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot update employees")
    
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    employee = await db.employees.find_one({
        "company_id": company_id,
        "chapa": employee_chapa
    }, {"_id": 0})
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    update_dict = {k: v for k, v in employee_data.model_dump().items() if v is not None}
    
    if update_dict:
        await db.employees.update_one(
            {"company_id": company_id, "chapa": employee_chapa},
            {"$set": update_dict}
        )
    
    updated_employee = await db.employees.find_one({
        "company_id": company_id,
        "chapa": employee_chapa
    }, {"_id": 0})
    
    if isinstance(updated_employee.get('created_at'), str):
        updated_employee['created_at'] = datetime.fromisoformat(updated_employee['created_at'])
    
    return Employee(**updated_employee)


@api_router.post("/employees/{employee_chapa}/deactivate")
async def deactivate_employee(
    employee_chapa: str,
    current_user: dict = Depends(get_current_user)
):
    """Deactivate (demobilize) an employee"""
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot deactivate employees")
    
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    result = await db.employees.update_one(
        {"company_id": company_id, "chapa": employee_chapa},
        {"$set": {"active": False}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    return {"message": "Employee deactivated successfully"}


@api_router.post("/employees/{employee_chapa}/reactivate")
async def reactivate_employee(
    employee_chapa: str,
    current_user: dict = Depends(get_current_user)
):
    """Reactivate (remobilize) an employee - viewers can do this"""
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    result = await db.employees.update_one(
        {"company_id": company_id, "chapa": employee_chapa},
        {"$set": {"active": True}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    return {"message": "Employee reactivated successfully"}


# ============ SHIFT ROUTES ============

@api_router.post("/shifts", response_model=Shift)
async def create_shift(
    shift_data: ShiftCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a shift configuration"""
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot create shifts")
    
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    shift = Shift(
        company_id=company_id,
        **shift_data.model_dump()
    )
    
    shift_dict = shift.model_dump()
    shift_dict['created_at'] = shift_dict['created_at'].isoformat()
    
    await db.shifts.insert_one(shift_dict)
    
    return shift


@api_router.get("/shifts", response_model=List[Shift])
async def get_shifts(
    current_user: dict = Depends(get_current_user)
):
    """Get shifts for company"""
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    shifts = await db.shifts.find({"company_id": company_id}, {"_id": 0}).to_list(1000)
    for shift in shifts:
        if isinstance(shift.get('created_at'), str):
            shift['created_at'] = datetime.fromisoformat(shift['created_at'])
    
    return shifts


@api_router.put("/shifts/{shift_id}", response_model=Shift)
async def update_shift(
    shift_id: str,
    shift_data: ShiftUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update shift configuration (only admin)"""
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot update shifts")
    
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    shift = await db.shifts.find_one({"id": shift_id, "company_id": company_id}, {"_id": 0})
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
    
    update_dict = {k: v for k, v in shift_data.model_dump().items() if v is not None}
    
    if update_dict:
        await db.shifts.update_one(
            {"id": shift_id, "company_id": company_id},
            {"$set": update_dict}
        )
    
    updated_shift = await db.shifts.find_one({"id": shift_id, "company_id": company_id}, {"_id": 0})
    if isinstance(updated_shift.get('created_at'), str):
        updated_shift['created_at'] = datetime.fromisoformat(updated_shift['created_at'])
    
    return Shift(**updated_shift)


@api_router.delete("/shifts/{shift_id}")
async def delete_shift(
    shift_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a shift"""
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot delete shifts")
    
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    result = await db.shifts.delete_one({"id": shift_id, "company_id": company_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Shift not found")
    
    return {"message": "Shift deleted successfully"}


# ============ CSV IMPORT ROUTES ============

@api_router.post("/import/csv")
async def import_csv_attendance(
    file: UploadFile = File(...),
    date: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Import CSV attendance data
    Format: CHAPA, NOME, FUNCAO, BATIDA
    Logic:
    - If BATIDA has time, mark as P (DIA) or PN (NOITE) based on employee's turno
    - If BATIDA is empty, mark as FALTA
    """
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot import data")
    
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    content = await file.read()
    decoded_content = content.decode('utf-8')
    csv_reader = csv.DictReader(io.StringIO(decoded_content))
    
    results = {
        "processed": 0,
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "errors": []
    }
    
    for row in csv_reader:
        try:
            chapa = row.get('CHAPA', '').strip()
            nome = row.get('NOME', '').strip()
            funcao = row.get('FUNCAO', '').strip()
            batida = row.get('BATIDA', '').strip()
            
            if not chapa:
                continue
            
            employee = await db.employees.find_one({
                "company_id": company_id,
                "chapa": chapa
            }, {"_id": 0})
            
            if not employee:
                results["skipped"] += 1
                results["errors"].append(f"Employee {chapa} not found")
                continue
            
            if batida:
                if employee['turno'] == 'DIA':
                    status = 'P'
                else:
                    status = 'PN'
                hora_batida = batida
            else:
                status = 'FALTA'
                hora_batida = None
            
            existing_attendance = await db.attendance.find_one({
                "company_id": company_id,
                "employee_chapa": chapa,
                "date": date
            })
            
            if existing_attendance:
                await db.attendance.update_one(
                    {
                        "company_id": company_id,
                        "employee_chapa": chapa,
                        "date": date
                    },
                    {
                        "$set": {
                            "status": status,
                            "hora_batida": hora_batida,
                            "registered_by": current_user['id'],
                            "registered_at": datetime.now(timezone.utc).isoformat()
                        }
                    }
                )
                results["updated"] += 1
            else:
                attendance = Attendance(
                    company_id=company_id,
                    employee_chapa=chapa,
                    date=date,
                    status=status,
                    hora_batida=hora_batida,
                    registered_by=current_user['id']
                )
                
                attendance_dict = attendance.model_dump()
                attendance_dict['registered_at'] = attendance_dict['registered_at'].isoformat()
                
                await db.attendance.insert_one(attendance_dict)
                results["created"] += 1
            
            results["processed"] += 1
            
        except Exception as e:
            results["errors"].append(f"Error processing row {chapa}: {str(e)}")
            continue
    
    return results


@api_router.post("/import/employees/csv")
async def import_csv_employees(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Import CSV employee registration data
    Format: chapa, nome, funcao, data_admissao
    Default values:
    - turno: DIA (can be edited later)
    - sindicato: 1 (can be edited later)
    - grupo: empty (can be edited later)
    - mo: empty (can be edited later)
    - primeiro_acesso: empty (can be edited later)
    """
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot import data")
    
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    content = await file.read()
    decoded_content = content.decode('utf-8')
    csv_reader = csv.DictReader(io.StringIO(decoded_content))
    
    results = {
        "processed": 0,
        "created": 0,
        "updated": 0,
        "errors": []
    }
    
    for row in csv_reader:
        try:
            chapa = row.get('chapa', '').strip()
            nome = row.get('nome', '').strip()
            funcao = row.get('funcao', '').strip()
            data_admissao = row.get('data_admissao', '').strip()
            
            if not chapa or not nome or not funcao:
                results["errors"].append(f"Missing required fields for row: {row}")
                continue
            
            # Check if employee already exists
            existing_employee = await db.employees.find_one({
                "company_id": company_id,
                "chapa": chapa
            })
            
            employee_data = {
                "chapa": chapa,
                "nome": nome,
                "funcao": funcao,
                "turno": "DIA",  # Default to DIA
                "grupo": "",  # Empty, to be edited later
                "mo": "",  # Empty, to be edited later
                "admissao": data_admissao if data_admissao else None,
                "sindicato": "1",  # Default to 1
                "primeiro_acesso": "",  # Empty, to be edited later
                "active": True,
                "company_id": company_id
            }
            
            if existing_employee:
                # Update existing employee
                await db.employees.update_one(
                    {"company_id": company_id, "chapa": chapa},
                    {"$set": employee_data}
                )
                results["updated"] += 1
            else:
                # Create new employee
                employee_data["id"] = str(uuid4())
                employee_data["created_at"] = datetime.now(timezone.utc).isoformat()
                await db.employees.insert_one(employee_data)
                results["created"] += 1
            
            results["processed"] += 1
            
        except Exception as e:
            results["errors"].append(f"Error processing row {row.get('chapa', 'unknown')}: {str(e)}")
            continue
    
    return results


# ============ ATTENDANCE ROUTES ============

@api_router.post("/attendance", response_model=Attendance)
async def create_attendance(
    attendance_data: AttendanceCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create or update attendance record"""
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot create attendance records")
    
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    employee = await db.employees.find_one({
        "company_id": company_id,
        "chapa": attendance_data.employee_chapa
    })
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    existing = await db.attendance.find_one({
        "company_id": company_id,
        "employee_chapa": attendance_data.employee_chapa,
        "date": attendance_data.date
    })
    
    if existing:
        update_dict = attendance_data.model_dump()
        update_dict['registered_by'] = current_user['id']
        update_dict['registered_at'] = datetime.now(timezone.utc).isoformat()
        
        await db.attendance.update_one(
            {
                "company_id": company_id,
                "employee_chapa": attendance_data.employee_chapa,
                "date": attendance_data.date
            },
            {"$set": update_dict}
        )
        
        updated = await db.attendance.find_one({
            "company_id": company_id,
            "employee_chapa": attendance_data.employee_chapa,
            "date": attendance_data.date
        }, {"_id": 0})
        
        if isinstance(updated.get('registered_at'), str):
            updated['registered_at'] = datetime.fromisoformat(updated['registered_at'])
        
        return Attendance(**updated)
    else:
        attendance = Attendance(
            company_id=company_id,
            registered_by=current_user['id'],
            **attendance_data.model_dump()
        )
        
        attendance_dict = attendance.model_dump()
        attendance_dict['registered_at'] = attendance_dict['registered_at'].isoformat()
        
        await db.attendance.insert_one(attendance_dict)
        
        return attendance


@api_router.get("/attendance", response_model=List[Attendance])
async def get_attendance(
    date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    employee_chapa: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get attendance records"""
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    query = {"company_id": company_id}
    
    if date:
        query["date"] = date
    elif start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    
    if employee_chapa:
        query["employee_chapa"] = employee_chapa
    
    records = await db.attendance.find(query, {"_id": 0}).to_list(10000)
    for record in records:
        if isinstance(record.get('registered_at'), str):
            record['registered_at'] = datetime.fromisoformat(record['registered_at'])
    
    return records


@api_router.put("/attendance/{attendance_id}", response_model=Attendance)
async def update_attendance(
    attendance_id: str,
    attendance_data: AttendanceUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update attendance record"""
    if current_user['role'] == 'company_viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot update attendance records")
    
    company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required")
    
    attendance = await db.attendance.find_one({
        "id": attendance_id,
        "company_id": company_id
    }, {"_id": 0})
    
    if not attendance:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    
    update_dict = {k: v for k, v in attendance_data.model_dump().items() if v is not None}
    update_dict['registered_by'] = current_user['id']
    update_dict['registered_at'] = datetime.now(timezone.utc).isoformat()
    
    if update_dict:
        await db.attendance.update_one(
            {"id": attendance_id, "company_id": company_id},
            {"$set": update_dict}
        )
    
    updated = await db.attendance.find_one({
        "id": attendance_id,
        "company_id": company_id
    }, {"_id": 0})
    
    if isinstance(updated.get('registered_at'), str):
        updated['registered_at'] = datetime.fromisoformat(updated['registered_at'])
    
    return Attendance(**updated)


# ============ EXPORT ROUTES ============

@api_router.get("/export/employees/excel")
async def export_employees_excel(
    turno: Optional[str] = None,
    active: Optional[bool] = True,
    current_user: dict = Depends(get_current_user)
):
    """Export employees to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        company_id = current_user['company_id'] if current_user['role'] != 'super_admin' else None
        if not company_id:
            raise HTTPException(status_code=400, detail="Company ID required")
        
        query = {"company_id": company_id}
        
        # Aplicar filtros
        if active is not None:
            query["active"] = active
        if turno:
            query["turno"] = turno
        
        employees = await db.employees.find(query, {"_id": 0}).to_list(10000)
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Efetivo {turno if turno else 'Todos'}"
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Cabeçalhos
        headers = ["Chapa", "Nome", "Função", "Turno", "Grupo", "M.O", "Admissão", "Sindicato"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        for row, emp in enumerate(employees, 2):
            ws.cell(row=row, column=1, value=emp['chapa'])
            ws.cell(row=row, column=2, value=emp['nome'])
            ws.cell(row=row, column=3, value=emp['funcao'])
            ws.cell(row=row, column=4, value=emp['turno'])
            ws.cell(row=row, column=5, value=emp['grupo'])
            ws.cell(row=row, column=6, value=emp['mo'])
            ws.cell(row=row, column=7, value=emp.get('admissao', ''))
            ws.cell(row=row, column=8, value=emp.get('sindicato', ''))
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from fastapi.responses import StreamingResponse
        
        # Gerar nome do arquivo baseado nos filtros
        if not active:
            filename_suffix = "desmobilizados"
        elif turno:
            filename_suffix = f"turno_{turno.lower()}"
        else:
            filename_suffix = "todos"
        
        filename = f"efetivo_{filename_suffix}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error exporting: {str(e)}")


# ============ PUBLIC VIEW ROUTES (Com validação de company_code) ============

@api_router.get("/public/employees")
async def get_public_employees(company_code: str):
    """Get employees for public view (requires company code)"""
    # Verificar se código da empresa existe
    company = await db.companies.find_one({"company_code": company_code}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Invalid company code")
    
    if not company.get('active', True):
        raise HTTPException(status_code=403, detail="Company account is inactive")
    
    # Buscar colaboradores ativos da empresa
    employees = await db.employees.find({
        "company_id": company['id'],
        "active": True
    }, {"_id": 0}).to_list(10000)
    
    for emp in employees:
        if isinstance(emp.get('created_at'), str):
            emp['created_at'] = datetime.fromisoformat(emp['created_at'])
    
    return employees


@api_router.get("/public/attendance")
async def get_public_attendance(
    company_code: str,
    date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Get attendance for public view (requires company code)"""
    # Verificar se código da empresa existe
    company = await db.companies.find_one({"company_code": company_code}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Invalid company code")
    
    if not company.get('active', True):
        raise HTTPException(status_code=403, detail="Company account is inactive")
    
    query = {"company_id": company['id']}
    
    if date:
        query["date"] = date
    elif start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    
    records = await db.attendance.find(query, {"_id": 0}).to_list(10000)
    for record in records:
        if isinstance(record.get('registered_at'), str):
            record['registered_at'] = datetime.fromisoformat(record['registered_at'])
    
    return records


@api_router.get("/public/export/excel")
async def export_public_excel(company_code: str, date: str, turno: Optional[str] = None):
    """Export public view to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # Verificar código da empresa
        company = await db.companies.find_one({"company_code": company_code}, {"_id": 0})
        if not company:
            raise HTTPException(status_code=404, detail="Invalid company code")
        
        # Buscar colaboradores e presença
        employees = await db.employees.find({"company_id": company['id'], "active": True}, {"_id": 0}).to_list(10000)
        attendance = await db.attendance.find({"company_id": company['id'], "date": date}, {"_id": 0}).to_list(10000)
        
        # Criar workbook com múltiplas abas
        wb = openpyxl.Workbook()
        
        # Remover aba padrão
        wb.remove(wb.active)
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Definir quais turnos exportar baseado no filtro
        if turno:
            # Exportar apenas o turno específico
            turnos_to_export = [turno]
        else:
            # Exportar ambos os turnos (comportamento padrão)
            turnos_to_export = ["DIA", "NOITE"]
        
        # Criar aba para cada turno
        for turno_atual in turnos_to_export:
            ws = wb.create_sheet(title=f"BASE {turno_atual}")
            
            # Cabeçalhos
            headers = ["Chapa", "Nome", "Função", "Grupo", "M.O", "Hora Batida", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Filtrar colaboradores por turno
            turno_employees = [e for e in employees if e['turno'] == turno_atual]
            
            # Dados
            for row, emp in enumerate(turno_employees, 2):
                att = next((a for a in attendance if a['employee_chapa'] == emp['chapa']), None)
                status = att['status'] if att else 'N/R'
                hora = att.get('hora_batida', '') if att else ''
                
                ws.cell(row=row, column=1, value=emp['chapa'])
                ws.cell(row=row, column=2, value=emp['nome'])
                ws.cell(row=row, column=3, value=emp['funcao'])
                ws.cell(row=row, column=4, value=emp['grupo'])
                ws.cell(row=row, column=5, value=emp['mo'])
                ws.cell(row=row, column=6, value=hora)
                ws.cell(row=row, column=7, value=status)
            
            # Ajustar largura
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Salvar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from fastapi.responses import StreamingResponse
        
        # Nome do arquivo baseado no filtro
        if turno:
            filename = f"efetivo_{company['name'].replace(' ', '_')}_{turno}_{date}.xlsx"
        else:
            filename = f"efetivo_{company['name'].replace(' ', '_')}_{date}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error exporting: {str(e)}")


# ============ PAYMENT SETTINGS ROUTES ============

from payment_config import (
    PaymentSettings, 
    get_payment_settings, 
    save_payment_settings,
    is_payment_configured,
    get_plan_price,
    calculate_subscription_end_date
)

@api_router.get("/settings/payment", response_model=PaymentSettings)
async def get_payment_config(current_user: dict = Depends(get_current_user)):
    """Get payment configuration (Super Admin only)"""
    if current_user['role'] != 'super_admin':
        raise HTTPException(status_code=403, detail="Only super admin can access settings")
    
    return get_payment_settings()

@api_router.post("/settings/payment")
async def save_payment_config(
    settings: PaymentSettings,
    current_user: dict = Depends(get_current_user)
):
    """Save payment configuration (Super Admin only)"""
    if current_user['role'] != 'super_admin':
        raise HTTPException(status_code=403, detail="Only super admin can access settings")
    
    success = save_payment_settings(settings)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to save settings")
    
    return {"message": "Settings saved successfully"}

@api_router.get("/settings/payment/status")
async def get_payment_status():
    """Get payment system status (public endpoint)"""
    return {
        "configured": is_payment_configured(),
        "message": "Payment system is configured" if is_payment_configured() else "Payment system not configured"
    }

@api_router.get("/settings/plans")
async def get_available_plans():
    """Get available subscription plans with prices"""
    settings = get_payment_settings()
    
    plans = []
    if settings.monthly_price:
        plans.append({
            "id": "monthly",
            "name": "Mensal",
            "duration": "1 mês",
            "price": float(settings.monthly_price),
            "description": "Renovação mensal automática"
        })
    
    if settings.semiannual_price:
        plans.append({
            "id": "semiannual",
            "name": "Semestral",
            "duration": "6 meses",
            "price": float(settings.semiannual_price),
            "description": "Economia de longo prazo"
        })
    
    if settings.annual_price:
        plans.append({
            "id": "annual",
            "name": "Anual",
            "duration": "12 meses",
            "price": float(settings.annual_price),
            "description": "Melhor custo-benefício"
        })
    
    return {
        "plans": plans,
        "configured": len(plans) > 0
    }


# ============ PAYMENT PROCESSING ROUTES ============

from pydantic import BaseModel

class LicensePaymentRequest(BaseModel):
    company_id: str
    plan_type: str  # monthly, semiannual, annual
    payment_method: str = "credit_card"
    card_token: Optional[str] = None  # Token do Mercado Pago
    payer_email: str
    payer_name: str

@api_router.post("/payments/license")
async def process_license_payment(
    payment_request: LicensePaymentRequest,
    current_user: dict = Depends(get_current_user)
):
    """Process license payment for a company"""
    
    # Verificar se pagamento está configurado
    if not is_payment_configured():
        raise HTTPException(
            status_code=503,
            detail="Payment system not configured. Please contact administrator."
        )
    
    # Buscar empresa
    company = await db.companies.find_one(
        {"id": payment_request.company_id},
        {"_id": 0}
    )
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Verificar permissões (Super Admin ou Admin da própria empresa)
    if current_user['role'] != 'super_admin':
        if current_user['company_id'] != payment_request.company_id:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Obter preço do plano
    plan_price = get_plan_price(payment_request.plan_type)
    if plan_price == 0:
        raise HTTPException(
            status_code=400,
            detail=f"Plan {payment_request.plan_type} not available or not configured"
        )
    
    # Calcular data de expiração
    new_expiration_date = calculate_subscription_end_date(payment_request.plan_type)
    
    # TODO: Integrar com Mercado Pago SDK quando credenciais estiverem configuradas
    # Por enquanto, simular pagamento aprovado
    settings = get_payment_settings()
    
    if settings.mp_access_token and payment_request.card_token:
        # Aqui seria a chamada real ao Mercado Pago
        # import mercadopago
        # sdk = mercadopago.SDK(settings.mp_access_token)
        # result = sdk.payment().create({...})
        
        # Por enquanto, simular aprovação
        payment_approved = True
    else:
        # Modo simulação - aprovar automaticamente
        payment_approved = True
    
    if payment_approved:
        # Atualizar empresa com nova data de expiração
        await db.companies.update_one(
            {"id": payment_request.company_id},
            {
                "$set": {
                    "subscription_expires_at": new_expiration_date.isoformat(),
                    "subscription_type": payment_request.plan_type,
                    "active": True
                }
            }
        )
        
        # Criar registro de pagamento
        payment_record = {
            "id": str(uuid4()),
            "company_id": payment_request.company_id,
            "company_name": company.get("name"),
            "plan_type": payment_request.plan_type,
            "amount": plan_price,
            "status": "approved",
            "payer_email": payment_request.payer_email,
            "payer_name": payment_request.payer_name,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "expires_at": new_expiration_date.isoformat()
        }
        
        await db.payments.insert_one(payment_record)
        
        return {
            "success": True,
            "message": "Payment processed successfully",
            "payment_id": payment_record["id"],
            "new_expiration_date": new_expiration_date.isoformat(),
            "plan_type": payment_request.plan_type,
            "amount": plan_price
        }
    else:
        raise HTTPException(
            status_code=400,
            detail="Payment was rejected. Please try again."
        )

@api_router.get("/payments/history/{company_id}")
async def get_payment_history(
    company_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get payment history for a company"""
    
    # Verificar permissões
    if current_user['role'] != 'super_admin':
        if current_user['company_id'] != company_id:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    payments = await db.payments.find(
        {"company_id": company_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return {"payments": payments}



# Mount the API router
app.include_router(api_router)


@app.get("/health")
async def health_check():
    return {"status": "ok", "service": "Employee Management System"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
