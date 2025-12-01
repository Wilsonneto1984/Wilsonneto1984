"""
Exportação de Efetivo Diário no formato personalizado
Layout baseado no modelo Petrobras
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any
from collections import defaultdict

def create_efetivo_diario_excel(
    company_name: str,
    contract_number: str,
    project_name: str,
    date: str,
    employees: List[Dict],
    attendance: List[Dict],
    subcontractors: List[Dict]
) -> Workbook:
    """
    Cria arquivo Excel com formato Efetivo Diário
    
    Estrutura:
    - Aba 1: Efetivo Diário (resumo consolidado)
    - Aba 2: Relação Nominal - DIA
    - Aba 3: Relação Nominal - NOITE
    """
    wb = Workbook()
    
    # ============ ABA 1: EFETIVO DIÁRIO ============
    ws_efetivo = wb.active
    ws_efetivo.title = "Efetivo Diário"
    
    # Criar cabeçalho
    create_header(ws_efetivo, company_name, contract_number, project_name, date)
    
    # Consolidar dados por função
    mo_indireta, mo_direta = consolidate_by_function(employees, attendance, date)
    
    # Criar seção Mão-de-Obra INDIRETA e DIRETA (lado a lado)
    current_row = create_mo_sections(ws_efetivo, mo_indireta, mo_direta, start_row=7)
    
    # Criar seção Mão-de-Obra SUBCONTRATADA
    current_row = create_subcontracted_section(ws_efetivo, subcontractors, start_row=current_row + 2)
    
    # Criar rodapé com definições e totais
    create_footer(ws_efetivo, mo_indireta, mo_direta, subcontractors, start_row=current_row + 2)
    
    # Ajustar largura das colunas
    adjust_column_widths(ws_efetivo)
    
    # ============ ABA 2: RELAÇÃO NOMINAL - DIA ============
    ws_dia = wb.create_sheet("Relação Nominal - DIA")
    create_nominal_sheet(ws_dia, employees, attendance, date, "DIA")
    
    # ============ ABA 3: RELAÇÃO NOMINAL - NOITE ============
    ws_noite = wb.create_sheet("Relação Nominal - NOITE")
    create_nominal_sheet(ws_noite, employees, attendance, date, "NOITE")
    
    return wb

def create_header(ws, company_name, contract_number, project_name, date):
    """Cria o cabeçalho do relatório"""
    # Título
    ws['A1'] = 'Cliente:'
    ws['B1'] = company_name
    ws['A2'] = 'Contrato nº:'
    ws['B2'] = contract_number or 'N/A'
    ws['A3'] = 'Obra:'
    ws['B3'] = project_name
    
    # Data
    date_obj = datetime.strptime(date, '%Y-%m-%d')
    ws['G1'] = 'Efetivo Dia:'
    ws['H1'] = date_obj.strftime('%d/%m/%Y')
    ws['G2'] = 'Dia da Semana:'
    ws['H2'] = date_obj.strftime('%A')
    
    # Estilo do cabeçalho
    for row in range(1, 4):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if col == 1 or col == 7:  # Labels
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Linha separadora
    ws.row_dimensions[5].height = 3

def consolidate_by_function(employees: List[Dict], attendance: List[Dict], date: str) -> tuple:
    """
    Consolida funcionários por função (Mão-de-Obra Indireta e Direta)
    """
    # Criar índice de presença por chapa
    attendance_map = {a['employee_chapa']: a for a in attendance if a['date'] == date}
    
    # Dicionários para consolidação
    mo_indireta = defaultdict(lambda: {'efetivo_geral': 0, 'ausencias': 0, 'presentes': 0})
    mo_direta = defaultdict(lambda: {'efetivo_geral': 0, 'ausencias': 0, 'presentes': 0})
    
    for emp in employees:
        if not emp.get('active'):
            continue
        
        funcao = emp.get('funcao', 'Não Informado')
        mo_type = emp.get('mo', '').upper()
        
        # Determinar se é MOI ou MOD
        if mo_type == 'M.O.I':
            target = mo_indireta
        else:
            target = mo_direta
        
        # Contar efetivo geral
        target[funcao]['efetivo_geral'] += 1
        
        # Verificar presença
        att = attendance_map.get(emp['chapa'])
        if att:
            status = att.get('status', 'FALTA')
            if status in ['P', 'PN']:
                target[funcao]['presentes'] += 1
            else:
                target[funcao]['ausencias'] += 1
        else:
            target[funcao]['ausencias'] += 1
    
    return dict(mo_indireta), dict(mo_direta)

def create_mo_sections(ws, mo_indireta, mo_direta, start_row):
    """
    Cria seções de Mão-de-Obra INDIRETA e DIRETA lado a lado
    """
    row = start_row
    
    # Título das seções
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'Mão-de-Obra INDIRETA'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    ws.merge_cells(f'F{row}:I{row}')
    ws[f'F{row}'] = 'Mão-de-Obra DIRETA'
    ws[f'F{row}'].font = Font(bold=True, size=12)
    ws[f'F{row}'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    row += 1
    
    # Cabeçalhos das colunas
    headers = ['Grupo', 'Função', 'Efetivo Geral', 'Ausências', 'Presentes']
    
    # MOI headers (colunas A-E)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # MOD headers (colunas F-J)
    for col_idx, header in enumerate(headers, start=6):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    row += 1
    start_data_row = row
    
    # Preencher dados MOI e MOD
    max_rows = max(len(mo_indireta), len(mo_direta))
    
    moi_items = sorted(mo_indireta.items())
    mod_items = sorted(mo_direta.items())
    
    for idx in range(max_rows):
        # MOI
        if idx < len(moi_items):
            funcao, data = moi_items[idx]
            ws.cell(row=row, column=1, value=idx + 1)  # Grupo
            ws.cell(row=row, column=2, value=funcao)
            ws.cell(row=row, column=3, value=data['efetivo_geral'])
            ws.cell(row=row, column=4, value=data['ausencias'])
            ws.cell(row=row, column=5, value=data['presentes'])
        
        # MOD
        if idx < len(mod_items):
            funcao, data = mod_items[idx]
            ws.cell(row=row, column=6, value=idx + 1)  # Grupo
            ws.cell(row=row, column=7, value=funcao)
            ws.cell(row=row, column=8, value=data['efetivo_geral'])
            ws.cell(row=row, column=9, value=data['ausencias'])
            ws.cell(row=row, column=10, value=data['presentes'])
        
        row += 1
    
    # Totais MOI
    total_moi_efetivo = sum(d['efetivo_geral'] for d in mo_indireta.values())
    total_moi_ausencias = sum(d['ausencias'] for d in mo_indireta.values())
    total_moi_presentes = sum(d['presentes'] for d in mo_indireta.values())
    
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = 'Total da Mão-de-Obra INDIRETA'
    ws[f'A{row}'].font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_moi_efetivo).font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_moi_ausencias).font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_moi_presentes).font = Font(bold=True)
    
    # Totais MOD
    total_mod_efetivo = sum(d['efetivo_geral'] for d in mo_direta.values())
    total_mod_ausencias = sum(d['ausencias'] for d in mo_direta.values())
    total_mod_presentes = sum(d['presentes'] for d in mo_direta.values())
    
    ws.merge_cells(f'F{row}:G{row}')
    ws[f'F{row}'] = 'Total da Mão-de-Obra DIRETA'
    ws[f'F{row}'].font = Font(bold=True)
    ws.cell(row=row, column=8, value=total_mod_efetivo).font = Font(bold=True)
    ws.cell(row=row, column=9, value=total_mod_ausencias).font = Font(bold=True)
    ws.cell(row=row, column=10, value=total_mod_presentes).font = Font(bold=True)
    
    return row

def create_subcontracted_section(ws, subcontractors, start_row):
    """Cria seção de Mão-de-Obra SUBCONTRATADA"""
    row = start_row
    
    # Título
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'Mão-de-Obra SUBCONTRATADA'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    row += 1
    
    # Cabeçalhos
    ws.cell(row=row, column=1, value='Função').font = Font(bold=True)
    ws.cell(row=row, column=2, value='Efetivo Geral').font = Font(bold=True)
    ws.cell(row=row, column=3, value='Presentes').font = Font(bold=True)
    
    row += 1
    
    # Dados das subcontratadas
    total_efetivo = 0
    total_presentes = 0
    
    for sub in subcontractors:
        if sub.get('active'):
            ws.cell(row=row, column=1, value=sub['name'])
            ws.cell(row=row, column=2, value=sub['employee_count'])
            ws.cell(row=row, column=3, value=sub['employee_count'])  # Assumindo todos presentes
            total_efetivo += sub['employee_count']
            total_presentes += sub['employee_count']
            row += 1
    
    # Total
    ws.cell(row=row, column=1, value='Total de SUBCONTRATADOS').font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_efetivo).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_presentes).font = Font(bold=True)
    
    return row

def create_footer(ws, mo_indireta, mo_direta, subcontractors, start_row):
    """Cria rodapé com definições e totais"""
    row = start_row
    
    # Definições
    definitions = [
        "EFETIVO GERAL: REFERE-SE A FUNCIONÁRIOS EFETIVAMENTE ADMITIDOS + FUNCIONÁRIOS APTOS PARA ADMISSÃO",
        "AUSÊNCIAS: QUALQUER TIPO DE AFASTAMENTO EX.: FALTAS / ATESTADOS ETC.",
        "FUNC EXT.: REFERE-SE AOS FUNCIONÁRIOS EFETIVAMENTE PRESENTES NA OBRA"
    ]
    
    for definition in definitions:
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = definition
        ws[f'A{row}'].font = Font(size=9)
        row += 1
    
    row += 1
    
    # Totais gerais
    total_efetivo_geral = (
        sum(d['efetivo_geral'] for d in mo_indireta.values()) +
        sum(d['efetivo_geral'] for d in mo_direta.values()) +
        sum(s['employee_count'] for s in subcontractors if s.get('active'))
    )
    
    total_ausencias = (
        sum(d['ausencias'] for d in mo_indireta.values()) +
        sum(d['ausencias'] for d in mo_direta.values())
    )
    
    total_presentes = total_efetivo_geral - total_ausencias
    
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = f'{total_efetivo_geral} - {total_ausencias} = {total_presentes}  Total de presentes'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

def create_nominal_sheet(ws, employees, attendance, date, shift):
    """Cria aba de Relação Nominal por turno"""
    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = f'Relação Nominal - Turno {shift}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Cabeçalhos
    headers = ['Chapa', 'Nome', 'Função', 'Grupo', 'MO', 'Status']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    # Filtrar funcionários do turno
    attendance_map = {a['employee_chapa']: a for a in attendance if a['date'] == date}
    
    row = 4
    for emp in sorted(employees, key=lambda x: x.get('nome', '')):
        if emp.get('turno') != shift or not emp.get('active'):
            continue
        
        att = attendance_map.get(emp['chapa'])
        status = att.get('status', 'FALTA') if att else 'FALTA'
        
        ws.cell(row=row, column=1, value=emp.get('chapa'))
        ws.cell(row=row, column=2, value=emp.get('nome'))
        ws.cell(row=row, column=3, value=emp.get('funcao'))
        ws.cell(row=row, column=4, value=emp.get('grupo', ''))
        ws.cell(row=row, column=5, value=emp.get('mo', ''))
        ws.cell(row=row, column=6, value=status)
        
        row += 1
    
    # Ajustar largura
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

def adjust_column_widths(ws):
    """Ajusta largura das colunas"""
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
